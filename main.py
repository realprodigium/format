import os
import io
import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.requests import Request

app = FastAPI(title="Conciliador XLSX")

# Configurar plantillas
templates = Jinja2Templates(directory="templates")

@app.get("/")
async def read_index(request: Request):
    return templates.TemplateResponse(request=request, name="index.html")

def clean_currency(value):
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    
    # Convertir a cadena y limpiar espacios y símbolos
    s_value = str(value).replace('$', '').strip()
    if not s_value or s_value == '.00':
        return 0.0

    # Determinar si el separador decimal es coma o punto
    # Si hay ambos, el que esté más a la derecha suele ser el decimal
    last_dot = s_value.rfind('.')
    last_comma = s_value.rfind(',')
    
    if last_dot > last_comma:
        # Formato US: 1,234.56 -> eliminar comas, mantener punto
        s_value = s_value.replace(',', '')
    elif last_comma > last_dot:
        # Formato EU: 1.234,56 -> eliminar puntos, cambiar coma por punto
        s_value = s_value.replace('.', '').replace(',', '.')
    else:
        # Solo hay uno o ninguno
        s_value = s_value.replace(',', '.')

    try:
        return float(s_value)
    except ValueError:
        return 0.0

@app.post("/api/process")
async def process_file(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Formato de archivo no soportado")

    try:
        # Leer el archivo excel completo inicialmente sin cabecera para encontrarla
        contents = await file.read()
        df_raw = pd.read_excel(io.BytesIO(contents), header=None)
        
        # Buscar la fila de cabecera
        header_row_index = 0
        target_keywords = ['fecha', 'descripcion', 'valor', 'saldo']
        for i, row in df_raw.iterrows():
            row_values = [str(x).lower().strip() for x in row.values if not pd.isna(x)]
            matches = sum(1 for kw in target_keywords if any(kw in rv for rv in row_values))
            if matches >= 2:
                header_row_index = i
                break
        
        # Volver a leer desde la fila encontrada
        df = pd.read_excel(io.BytesIO(contents), header=header_row_index)

        # 1. Limpiar columnas completamente vacías (como Unnamed: 6, 7)
        df = df.dropna(axis=1, how='all')

        # Normalizar nombres de columnas
        df.columns = [str(c).strip() for c in df.columns]
        
        # 2. Identificar y filtrar filas de basura (repetidas cabeceras o metadatos)
        # Identificar columnas de Valor y Saldo
        valor_cols = [c for c in df.columns if 'valor' in c.lower()]
        saldo_cols = [c for c in df.columns if 'saldo' in c.lower()]
        fecha_cols = [c for c in df.columns if 'fecha' in c.lower()]
        
        # Filtro de filas:
        if fecha_cols:
            fecha_col = fecha_cols[0]
            # Eliminar filas donde la fecha sea igual al nombre de la columna (cabeceras repetidas)
            df = df[df[fecha_col].astype(str).str.lower() != fecha_col.lower()]
            
            # Eliminar filas de metadatos comunes en extractos bancarios
            garbage_keywords = ['información', 'cliente:', 'dirección', 'desde', 'hasta', 'movimientos:', 'resumen:']
            mask = df[fecha_col].astype(str).str.lower().str.contains('|'.join(garbage_keywords), na=False)
            df = df[~mask]
            
            # Eliminar filas donde casi todo sea NaN (posibles separadores)
            df = df.dropna(subset=[fecha_col])

        # Limpiar columnas de dinero
        cols_to_fix = valor_cols + saldo_cols
        for col in cols_to_fix:
            df[col] = df[col].apply(clean_currency)

        # 3. Eliminar filas que quedaron en 0 en Valor y Saldo después de la limpieza y que no tienen descripción real
        # (Esto ayuda a eliminar filas de totales o subtotales que no queremos en la tabla de datos)
        if valor_cols and saldo_cols:
            desc_cols = [c for c in df.columns if 'descrip' in c.lower()]
            if desc_cols:
                desc_col = desc_cols[0]
                # Si el valor y saldo son 0 y la descripción parece metadato, borrar
                # (A veces hay transacciones de 0, pero suelen tener descripción clara)
                metadata_desc = ['desde', 'hasta', 'nro cuenta', 'cliente', 'sucursal']
                mask_meta = df[desc_col].astype(str).str.lower().str.contains('|'.join(metadata_desc), na=False)
                df = df[~mask_meta]

        # Exportar a Excel con formato de tabla
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data Procesada')
            
            workbook = writer.book
            worksheet = writer.sheets['Data Procesada']
            
            num_rows, num_cols = df.shape
            if num_rows > 0:
                from openpyxl.utils import get_column_letter
                last_col_letter = get_column_letter(num_cols)
                table_range = f"A1:{last_col_letter}{num_rows + 1}"
                
                from openpyxl.worksheet.table import Table, TableStyleInfo
                tab = Table(displayName="TablaProcesada", ref=table_range)
                
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                worksheet.add_table(tab)
                
                # Aplicar formato numérico explícito
                from openpyxl.styles import NamedStyle
                number_format = '#,##0.00'
                
                for i, col_name in enumerate(df.columns):
                    col_letter = get_column_letter(i + 1)
                    is_money = any(kw in col_name.lower() for kw in ['valor', 'saldo'])
                    
                    if is_money:
                        for cell in worksheet[col_letter][1:]:
                            cell.number_format = number_format
                    
                    # Ajustar ancho de columnas
                    data_max = 0
                    for val in df[col_name]:
                        val_str = str(val) if not pd.isna(val) else ""
                        if len(val_str) > data_max:
                            data_max = len(val_str)
                    
                    head_len = len(str(col_name))
                    column_length = max(data_max, head_len) + 4
                    worksheet.column_dimensions[col_letter].width = column_length
            else:
                # Si no hay datos, al menos poner un mensaje o dejar la hoja vacía
                worksheet.cell(row=1, column=1, value="No se encontraron transacciones en el formato esperado.")

        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=procesado_{file.filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
